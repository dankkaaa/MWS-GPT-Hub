import hashlib
import mimetypes
import os
import re
import subprocess
import zipfile
from functools import lru_cache
from html import unescape
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None

from llm_core.schemas import Attachment, ChatRequest, MemoryPacket, ToolResult
from llm_core.storage_backend import StorageBackend


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DB_PATH = Path(os.getenv("MTS_MEMORY_DB_PATH") or (REPO_ROOT / ".mts_memory" / "memory.db"))
URL_PATTERN = re.compile(r"https?://[^\s<>\"]+")
WORD_PATTERN = re.compile(r"\w+", flags=re.UNICODE)
RUS_STOPWORDS = {
    "это",
    "как",
    "что",
    "для",
    "или",
    "при",
    "под",
    "над",
    "они",
    "она",
    "его",
    "ему",
    "еще",
    "если",
    "когда",
    "тогда",
    "после",
    "перед",
    "через",
    "про",
    "если",
    "надо",
    "нужно",
    "сейчас",
    "этой",
    "этот",
    "этом",
    "того",
    "тебе",
    "меня",
    "мне",
}
EN_STOPWORDS = {
    "this",
    "that",
    "with",
    "from",
    "have",
    "into",
    "about",
    "your",
    "what",
    "when",
    "where",
    "there",
    "would",
    "could",
    "should",
    "need",
    "latest",
    "please",
}
TASK_PREFIXES = (
    "сделать",
    "добавить",
    "реализовать",
    "докрутить",
    "подключить",
    "собрать",
    "настроить",
    "вернуть",
    "заполнить",
    "построить",
    "описать",
    "ingest",
    "index",
    "implement",
    "wire",
    "build",
)
LOW_SIGNAL_MARKERS = {
    "да",
    "ага",
    "угу",
    "ок",
    "окей",
    "понял",
    "поняла",
    "хорошо",
    "ладно",
    "нет",
    "неа",
    "да хочу",
    "я хочу",
    "не так",
    "неверно",
}
FIRST_PERSON_PATTERNS = (
    re.compile(r"^\s*меня зовут\b", flags=re.IGNORECASE),
    re.compile(r"^\s*я\b", flags=re.IGNORECASE),
    re.compile(r"^\s*мы\b", flags=re.IGNORECASE),
    re.compile(r"\bя работаю\b", flags=re.IGNORECASE),
    re.compile(r"\bя делаю\b", flags=re.IGNORECASE),
    re.compile(r"\bя строю\b", flags=re.IGNORECASE),
    re.compile(r"\bмы делаем\b", flags=re.IGNORECASE),
    re.compile(r"\bнаш проект\b", flags=re.IGNORECASE),
)
PREFERENCE_PATTERNS = (
    re.compile(r"\bпредпоч", flags=re.IGNORECASE),
    re.compile(r"\bмне важно\b", flags=re.IGNORECASE),
    re.compile(r"\bважен\b", flags=re.IGNORECASE),
    re.compile(r"\bважна\b", flags=re.IGNORECASE),
    re.compile(r"\bне надо\b", flags=re.IGNORECASE),
    re.compile(r"\bне нужен\b", flags=re.IGNORECASE),
    re.compile(r"\bне нужна\b", flags=re.IGNORECASE),
    re.compile(r"\bбез воды\b", flags=re.IGNORECASE),
    re.compile(r"\bкоротко\b", flags=re.IGNORECASE),
    re.compile(r"\bподробно\b", flags=re.IGNORECASE),
    re.compile(r"\bstable\b", flags=re.IGNORECASE),
    re.compile(r"\bclean context\b", flags=re.IGNORECASE),
)


def clean_text(text: str) -> str:
    text = unescape(text or "")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def truncate(text: str, limit: int) -> str:
    text = clean_text(text)
    if len(text) <= limit:
        return text
    clipped = text[: max(0, limit - 1)].rstrip()
    return clipped + "…"


def normalize_key(text: str) -> str:
    return re.sub(r"\s+", " ", clean_text(text).lower())


def tokenize(text: str) -> list[str]:
    tokens: list[str] = []
    for token in WORD_PATTERN.findall(text.lower()):
        if len(token) < 3 or token.isdigit():
            continue
        if token in RUS_STOPWORDS or token in EN_STOPWORDS:
            continue
        tokens.append(token)
    return tokens


def stable_hash(*parts: str) -> str:
    digest = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()
    return digest[:16]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file_pointer:
        while True:
            chunk = file_pointer.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def extract_urls(text: str) -> list[str]:
    return [match.group(0).rstrip(".,)") for match in URL_PATTERN.finditer(text or "")]


def split_sentences(text: str) -> list[str]:
    prepared = clean_text(text)
    parts = re.split(r"(?<=[.!?])\s+|\n+", prepared)
    return [part.strip(" -•\t") for part in parts if part.strip(" -•\t")]


def looks_like_task(line: str) -> bool:
    lowered = line.lower().strip()
    return any(lowered.startswith(prefix) for prefix in TASK_PREFIXES) or any(
        marker in lowered
        for marker in (
            "нужно",
            "надо",
            "что нужно сделать",
            "todo",
            "task",
            "pipeline",
            "retrieval",
            "storage",
            "memory api",
        )
    )


def build_chunks(text: str, chunk_words: int = 180, overlap_words: int = 35) -> list[str]:
    words = clean_text(text).split()
    if not words:
        return []
    if len(words) <= chunk_words:
        return [" ".join(words)]

    chunks: list[str] = []
    step = max(1, chunk_words - overlap_words)
    for start in range(0, len(words), step):
        chunk = " ".join(words[start : start + chunk_words]).strip()
        if not chunk:
            continue
        if chunks and chunk == chunks[-1]:
            continue
        chunks.append(chunk)
        if start + chunk_words >= len(words):
            break
    return chunks


def is_low_signal_message(text: str) -> bool:
    normalized = " ".join(clean_text(text).lower().split())
    if not normalized:
        return True
    if normalized in LOW_SIGNAL_MARKERS:
        return True
    words = normalized.split()
    if len(words) <= 3 and any(word in LOW_SIGNAL_MARKERS for word in words):
        return True
    return False


def looks_like_question(text: str) -> bool:
    normalized = clean_text(text).lower()
    if "?" in normalized:
        return True
    return normalized.startswith(
        (
            "кто ",
            "что ",
            "как ",
            "какая ",
            "какой ",
            "какие ",
            "где ",
            "когда ",
            "почему ",
            "зачем ",
            "расскажи ",
            "скажи ",
        )
    )


def is_first_person_statement(text: str) -> bool:
    if looks_like_question(text):
        return False
    if len(clean_text(text).split()) < 3:
        return False
    return any(pattern.search(text) for pattern in FIRST_PERSON_PATTERNS)


def is_preference_statement(text: str) -> bool:
    if looks_like_question(text):
        return False
    return any(pattern.search(text) for pattern in PREFERENCE_PATTERNS)


def serialize_attachment(attachment: Attachment) -> dict[str, Any]:
    return {
        "type": attachment.type,
        "file_id": attachment.file_id,
        "url": attachment.url,
        "mime_type": attachment.mime_type,
        "metadata": attachment.metadata,
    }


def score_text(query: str, text: str, same_chat: bool = False) -> float:
    query_tokens = tokenize(query)
    text_tokens = set(tokenize(text))
    if not query_tokens:
        return 0.1 + (0.15 if same_chat else 0.0)

    overlap = sum(1 for token in query_tokens if token in text_tokens)
    if overlap == 0:
        normalized_query = normalize_key(query)
        if normalized_query and normalized_query in normalize_key(text):
            overlap = 1
    if overlap == 0:
        return 0.0
    score = overlap * 2.0
    if same_chat:
        score += 0.35
    return score


def pick_top_rows(
    rows: list[dict[str, Any]],
    query: str,
    text_key: str,
    limit: int,
    chat_id: str | None = None,
    fallback_if_no_match: bool = False,
) -> list[dict[str, Any]]:
    query_tokens = tokenize(query)
    scored: list[tuple[float, int, dict[str, Any]]] = []
    for index, row in enumerate(rows):
        same_chat = bool(chat_id and row.get("chat_id") == chat_id)
        score = score_text(query, row.get(text_key, ""), same_chat=same_chat)
        score += max(0.0, 0.03 - index * 0.001)
        scored.append((score, index, row))
    scored.sort(key=lambda item: (item[0], -item[1]), reverse=True)
    selected = [row for score, _, row in scored if score > 0][:limit]
    if selected:
        return selected
    if fallback_if_no_match or not query_tokens:
        return rows[:limit]
    return []


class FileParser:
    def parse(self, file_path: Path, mime_type: str | None = None) -> tuple[str, dict[str, Any]]:
        path = file_path.expanduser().resolve()
        guessed_mime = mime_type or mimetypes.guess_type(path.name)[0] or ""
        suffix = path.suffix.lower()

        if suffix in {".txt", ".md", ".py", ".json", ".csv", ".ts", ".js", ".tsx", ".jsx", ".yaml", ".yml"}:
            text = self._read_text(path)
            return text, {"parser": "plain_text", "mime_type": guessed_mime or "text/plain"}

        if suffix in {".html", ".htm", ".xml"}:
            text = self._read_html(path.read_text(encoding="utf-8", errors="replace"))
            return text, {"parser": "html", "mime_type": guessed_mime or "text/html"}

        if suffix == ".docx":
            text = self._read_docx(path)
            return text, {"parser": "docx", "mime_type": guessed_mime or "application/vnd.openxmlformats-officedocument.wordprocessingml.document"}

        if suffix == ".xlsx":
            text = self._read_xlsx(path)
            return text, {"parser": "xlsx", "mime_type": guessed_mime or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}

        if suffix == ".pdf":
            text = self._read_pdf(path)
            return text, {"parser": "pdf", "mime_type": guessed_mime or "application/pdf"}

        if guessed_mime.startswith("text/"):
            text = self._read_text(path)
            return text, {"parser": "plain_text", "mime_type": guessed_mime}

        text = self._read_text(path)
        return text, {"parser": "best_effort_text", "mime_type": guessed_mime}

    def _read_text(self, path: Path) -> str:
        return clean_text(path.read_text(encoding="utf-8", errors="replace"))

    def _read_html(self, html: str) -> str:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style", "noscript", "svg"]):
            tag.decompose()
        return clean_text(soup.get_text("\n", strip=True))

    def _read_docx(self, path: Path) -> str:
        with zipfile.ZipFile(path) as archive:
            raw_xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
        text = re.sub(r"</w:p>", "\n", raw_xml)
        text = re.sub(r"<[^>]+>", " ", text)
        return clean_text(text)

    def _read_xlsx(self, path: Path) -> str:
        if load_workbook is None:
            raise RuntimeError("openpyxl is not installed")
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        rows: list[str] = []
        for sheet in workbook.worksheets:
            rows.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
        workbook.close()
        return clean_text("\n".join(rows))

    def _read_pdf(self, path: Path) -> str:
        pdftotext_path = shutil_which("pdftotext")
        if pdftotext_path:
            result = subprocess.run(
                [pdftotext_path, "-q", "-layout", str(path), "-"],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.stdout.strip():
                return clean_text(result.stdout)

        strings_path = shutil_which("strings")
        if strings_path:
            result = subprocess.run(
                [strings_path, "-n", "6", str(path)],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.stdout.strip():
                return clean_text(result.stdout)

        raise RuntimeError("No PDF parser is available")


class UrlParser:
    def fetch(self, url: str) -> tuple[str, str, dict[str, Any]]:
        response = requests.get(
            url,
            timeout=15,
            headers={"User-Agent": "MTS-MVP/1.0 (+retrieval-backend)"},
        )
        response.raise_for_status()

        content_type = response.headers.get("content-type", "")
        title = ""
        if "html" in content_type or "<html" in response.text.lower():
            soup = BeautifulSoup(response.text, "html.parser")
            for tag in soup(["script", "style", "noscript", "svg", "iframe", "form"]):
                tag.decompose()
            title = clean_text(soup.title.get_text(" ", strip=True)) if soup.title else ""
            main_node = soup.find("main") or soup.find("article") or soup.body or soup
            text = clean_text(main_node.get_text("\n", strip=True))
        else:
            text = clean_text(response.text)
        return title, text, {"content_type": content_type, "status_code": response.status_code}


def shutil_which(name: str) -> str | None:
    for directory in os.getenv("PATH", "").split(os.pathsep):
        candidate = Path(directory) / name
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


class IngestionService:
    def __init__(
        self,
        storage: StorageBackend,
        file_parser: FileParser | None = None,
        url_parser: UrlParser | None = None,
    ) -> None:
        self.storage = storage
        self.file_parser = file_parser or FileParser()
        self.url_parser = url_parser or UrlParser()

    def ingest_file(
        self,
        user_id: str,
        chat_id: str,
        file_path: str,
        file_id: str | None = None,
        mime_type: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        try:
            path = Path(file_path).expanduser().resolve()
            resolved_file_id = file_id or f"file-{stable_hash(user_id, chat_id, path.name, str(path.stat().st_mtime_ns))}"
            file_sha = sha256_file(path)
            raw_text, parser_metadata = self.file_parser.parse(path, mime_type=mime_type)
            chunks = build_chunks(raw_text)
            status = "indexed" if chunks else "empty"
            merged_metadata = {**(metadata or {}), **parser_metadata, "chunk_count": len(chunks), "text_length": len(raw_text)}
            self.storage.upsert_file(
                file_id=resolved_file_id,
                user_id=user_id,
                chat_id=chat_id,
                file_name=path.name,
                mime_type=parser_metadata.get("mime_type", mime_type or ""),
                file_path=str(path),
                raw_text=raw_text,
                sha256=file_sha,
                status=status,
                metadata=merged_metadata,
            )
            chunk_metadata = [{"file_name": path.name, "source_ref": f"{resolved_file_id}:chunk-{index + 1}"} for index in range(len(chunks))]
            chunk_ids = self.storage.replace_chunks(
                source_type="file",
                source_id=resolved_file_id,
                user_id=user_id,
                chat_id=chat_id,
                chunks=chunks,
                metadata=chunk_metadata,
            )
            return {
                "status": status,
                "file_id": resolved_file_id,
                "file_name": path.name,
                "mime_type": parser_metadata.get("mime_type", mime_type or ""),
                "chunk_ids": chunk_ids,
                "metadata": merged_metadata,
            }
        except Exception as error:
            path = Path(file_path).expanduser()
            resolved_file_id = file_id or f"file-{stable_hash(user_id, chat_id, path.name, file_path)}"
            merged_metadata = {**(metadata or {}), "error": str(error)}
            self.storage.upsert_file(
                file_id=resolved_file_id,
                user_id=user_id,
                chat_id=chat_id,
                file_name=path.name,
                mime_type=mime_type or mimetypes.guess_type(path.name)[0] or "",
                file_path=str(path),
                raw_text="",
                sha256="",
                status="error",
                metadata=merged_metadata,
            )
            return {
                "status": "error",
                "file_id": resolved_file_id,
                "file_name": path.name,
                "chunk_ids": [],
                "metadata": merged_metadata,
            }

    def ingest_url(
        self,
        user_id: str,
        chat_id: str,
        url: str,
        metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        resolved_url = url.strip()
        link_id = f"url-{stable_hash(user_id, chat_id, resolved_url)}"
        try:
            title, cleaned_page, parser_metadata = self.url_parser.fetch(resolved_url)
            chunks = build_chunks(cleaned_page)
            status = "indexed" if chunks else "empty"
            merged_metadata = {**(metadata or {}), **parser_metadata, "chunk_count": len(chunks), "text_length": len(cleaned_page)}
            self.storage.upsert_link(
                link_id=link_id,
                user_id=user_id,
                chat_id=chat_id,
                url=resolved_url,
                title=title,
                cleaned_text=cleaned_page,
                status=status,
                metadata=merged_metadata,
            )
            chunk_metadata = [{"url": resolved_url, "title": title, "source_ref": f"{link_id}:chunk-{index + 1}"} for index in range(len(chunks))]
            chunk_ids = self.storage.replace_chunks(
                source_type="url",
                source_id=link_id,
                user_id=user_id,
                chat_id=chat_id,
                chunks=chunks,
                metadata=chunk_metadata,
            )
            return {
                "status": status,
                "link_id": link_id,
                "url": resolved_url,
                "title": title,
                "chunk_ids": chunk_ids,
                "metadata": merged_metadata,
            }
        except Exception as error:
            merged_metadata = {**(metadata or {}), "error": str(error)}
            self.storage.upsert_link(
                link_id=link_id,
                user_id=user_id,
                chat_id=chat_id,
                url=resolved_url,
                title="",
                cleaned_text="",
                status="error",
                metadata=merged_metadata,
            )
            return {
                "status": "error",
                "link_id": link_id,
                "url": resolved_url,
                "title": "",
                "chunk_ids": [],
                "metadata": merged_metadata,
            }


class RetrievalService:
    def __init__(self, storage: StorageBackend) -> None:
        self.storage = storage

    def search_chunks(
        self,
        user_id: str,
        chat_id: str,
        query: str,
        source_type: str | None = None,
        source_ids: list[str] | None = None,
        limit: int = 4,
        one_per_source: bool = False,
    ) -> list[dict[str, Any]]:
        rows = self.storage.list_chunks(
            user_id=user_id,
            chat_id=None,
            source_type=source_type,
            source_ids=source_ids,
            limit=300,
        )
        ranked = pick_top_rows(rows, query=query, text_key="text", limit=50, chat_id=chat_id)
        if not ranked and rows:
            ranked = pick_top_rows(
                rows,
                query=query,
                text_key="text",
                limit=50,
                chat_id=chat_id,
                fallback_if_no_match=True,
            )
        if not one_per_source:
            return ranked[:limit]

        selected: list[dict[str, Any]] = []
        seen_sources: set[str] = set()
        for row in ranked:
            source_id = row["source_id"]
            if source_id in seen_sources:
                continue
            seen_sources.add(source_id)
            selected.append(row)
            if len(selected) >= limit:
                break
        return selected

    def _source_title(self, row: dict[str, Any]) -> str:
        if row["source_type"] == "file":
            file_record = self.storage.get_file(row["source_id"])
            if file_record:
                return file_record.get("file_name") or row["source_id"]
        if row["source_type"] == "url":
            link_record = self.storage.get_link_by_id(row["source_id"])
            if link_record:
                return link_record.get("title") or link_record.get("url") or row["source_id"]
        return row["source_id"]

    def format_memory_docs(self, chunks: list[dict[str, Any]], limit: int = 3) -> list[str]:
        lines: list[str] = []
        for chunk in chunks[:limit]:
            source_ref = f"{chunk['source_id']}:chunk-{chunk['ordinal'] + 1}"
            label = self._source_title(chunk)
            lines.append(f"[{source_ref}] {label}: {truncate(chunk['text'], 220)}")
        return lines

    def pack_tool_context(
        self,
        chunks: list[dict[str, Any]],
        max_chars: int = 1800,
    ) -> tuple[str, list[str]]:
        sections: list[str] = []
        sources: list[str] = []
        total = 0
        for chunk in chunks:
            source_ref = f"{chunk['source_id']}:chunk-{chunk['ordinal'] + 1}"
            snippet = truncate(chunk["text"], 500)
            block = f"[Source: {source_ref}]\n{snippet}"
            projected = total + len(block) + 2
            if sections and projected > max_chars:
                break
            sections.append(block)
            sources.append(source_ref)
            total = projected
        return "\n\n".join(sections), sources


class MemoryService:
    def __init__(self, storage: StorageBackend, retrieval: RetrievalService) -> None:
        self.storage = storage
        self.retrieval = retrieval

    def fetch_memory_context(
        self,
        user_id: str,
        chat_id: str,
        message: str,
        task_type: str = "text_chat",
        attachments: list[Attachment] | None = None,
    ) -> dict[str, Any]:
        low_signal_message = is_low_signal_message(message)
        summary_row = self.storage.get_chat_summary(chat_id)
        chat_summary = summary_row["summary"] if summary_row else self._build_summary(chat_id)
        facts = [] if low_signal_message else pick_top_rows(
            self.storage.list_user_facts(user_id),
            message,
            "fact",
            limit=4,
            chat_id=chat_id,
        )
        preferences = [] if low_signal_message else pick_top_rows(
            self.storage.list_preferences(user_id),
            message,
            "preference",
            limit=4,
            chat_id=chat_id,
        )
        tasks = [] if low_signal_message else pick_top_rows(
            self.storage.list_tasks(user_id),
            message,
            "task",
            limit=4,
            chat_id=chat_id,
        )

        preferred_source_ids = [
            attachment.file_id
            for attachment in (attachments or [])
            if attachment.type == "file" and attachment.file_id
        ]
        docs = [] if low_signal_message and not preferred_source_ids else self.retrieval.search_chunks(
            user_id=user_id,
            chat_id=chat_id,
            query=message,
            source_type="file" if preferred_source_ids else None,
            source_ids=preferred_source_ids or None,
            limit=3,
            one_per_source=True,
        )
        if not docs and preferred_source_ids:
            docs = self.retrieval.search_chunks(
                user_id=user_id,
                chat_id=chat_id,
                query=message,
                limit=3,
                one_per_source=True,
            )

        return {
            "chat_summary": truncate(chat_summary, 500),
            "user_facts": [truncate(item["fact"], 160) for item in facts[:4]],
            "preferences": [truncate(item["preference"], 160) for item in preferences[:4]],
            "relevant_docs": self.retrieval.format_memory_docs(docs, limit=3),
            "ongoing_tasks": [truncate(item["task"], 160) for item in tasks[:4] if item.get("status", "open") == "open"],
            "task_type": task_type,
        }

    def write_memory(
        self,
        user_id: str,
        chat_id: str,
        message: str,
        answer: str,
        task_type: str = "text_chat",
        attachments: list[Attachment] | None = None,
        answer_metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        message_metadata = {
            "attachments": [serialize_attachment(item) for item in attachments or []],
            "task_type": task_type,
        }
        self.storage.add_message(
            user_id=user_id,
            chat_id=chat_id,
            role="user",
            content=message,
            task_type=task_type,
            metadata=message_metadata,
        )
        self.storage.add_message(
            user_id=user_id,
            chat_id=chat_id,
            role="assistant",
            content=answer,
            task_type=task_type,
            metadata={"task_type": task_type, **(answer_metadata or {})},
        )

        summary = self._build_summary(chat_id)
        self.storage.upsert_chat_summary(user_id=user_id, chat_id=chat_id, summary=summary)

        written_facts = self._store_user_facts(user_id=user_id, chat_id=chat_id, message=message)
        written_preferences = self._store_preferences(user_id=user_id, chat_id=chat_id, message=message)
        written_tasks = self._store_tasks(user_id=user_id, chat_id=chat_id, message=message, task_type=task_type)

        return {
            "stored": True,
            "chat_summary": summary,
            "user_facts_written": written_facts,
            "preferences_written": written_preferences,
            "ongoing_tasks_written": written_tasks,
        }

    def _build_summary(self, chat_id: str) -> str:
        messages = self.storage.list_recent_messages(chat_id, limit=6)
        if not messages:
            return ""
        snippets: list[str] = []
        for message in messages:
            role = "Пользователь" if message["role"] == "user" else "Ассистент"
            snippets.append(f"{role}: {truncate(message['content'], 110)}")
        return truncate("Недавний контекст: " + " | ".join(snippets), 500)

    def _store_user_facts(self, user_id: str, chat_id: str, message: str) -> list[str]:
        stored: list[str] = []
        for sentence in split_sentences(message):
            if not is_first_person_statement(sentence):
                continue
            fact = truncate(sentence, 180)
            if len(fact) < 12:
                continue
            self.storage.upsert_user_fact(
                user_id=user_id,
                fact=fact,
                normalized_fact=normalize_key(fact),
                source_chat_id=chat_id,
                metadata={"source": "message"},
            )
            stored.append(fact)
        return stored[:4]

    def _store_preferences(self, user_id: str, chat_id: str, message: str) -> list[str]:
        stored: list[str] = []
        for sentence in split_sentences(message):
            if not is_preference_statement(sentence):
                continue
            preference = truncate(sentence, 180)
            self.storage.upsert_preference(
                user_id=user_id,
                preference=preference,
                normalized_preference=normalize_key(preference),
                source_chat_id=chat_id,
                metadata={"source": "message"},
            )
            stored.append(preference)
        return stored[:4]

    def _store_tasks(self, user_id: str, chat_id: str, message: str, task_type: str) -> list[str]:
        stored: list[str] = []
        lines = [line.strip(" -•\t") for line in clean_text(message).splitlines() if line.strip(" -•\t")]
        candidates = [line for line in lines if looks_like_task(line)]

        for candidate in candidates[:6]:
            task = truncate(candidate, 180)
            self.storage.upsert_task(
                user_id=user_id,
                chat_id=chat_id,
                task=task,
                normalized_task=normalize_key(task),
                status="open",
                metadata={"source": "message", "task_type": task_type},
            )
            stored.append(task)
        return stored[:4]


class ToolBackend:
    def __init__(
        self,
        storage: StorageBackend,
        ingestion: IngestionService,
        retrieval: RetrievalService,
    ) -> None:
        self.storage = storage
        self.ingestion = ingestion
        self.retrieval = retrieval

    def retrieve_doc_context(self, request: ChatRequest) -> ToolResult:
        file_ids: list[str] = []
        for attachment in request.attachments:
            if attachment.type != "file":
                continue
            resolved_file_id = self._ensure_file_attachment_ingested(
                user_id=request.user_id,
                chat_id=request.chat_id,
                attachment=attachment,
            )
            if resolved_file_id:
                file_ids.append(resolved_file_id)

        chunks = self.retrieval.search_chunks(
            user_id=request.user_id,
            chat_id=request.chat_id,
            query=request.message,
            source_type="file",
            source_ids=file_ids or None,
            limit=4,
        )
        content, sources = self.retrieval.pack_tool_context(chunks)
        return ToolResult(
            name="retrieve_doc_context",
            content=content,
            metadata={"sources": sources, "file_ids": file_ids},
        )

    def parse_url(self, request: ChatRequest) -> ToolResult:
        url = self._resolve_url(request)
        if not url:
            return ToolResult(name="parse_url", content="", metadata={"error": "url_not_found"})

        link = self.storage.get_link_by_url(request.user_id, request.chat_id, url)
        if link is None or link.get("status") != "indexed":
            self.ingestion.ingest_url(user_id=request.user_id, chat_id=request.chat_id, url=url)
            link = self.storage.get_link_by_url(request.user_id, request.chat_id, url)

        if link is None:
            return ToolResult(name="parse_url", content="", metadata={"url": url, "error": "ingestion_failed"})

        chunks = self.retrieval.search_chunks(
            user_id=request.user_id,
            chat_id=request.chat_id,
            query=request.message,
            source_type="url",
            source_ids=[link["link_id"]],
            limit=3,
        )
        if not chunks:
            chunks = self.storage.get_chunks_for_source("url", link["link_id"], limit=3)

        content, sources = self.retrieval.pack_tool_context(chunks, max_chars=2200)
        return ToolResult(
            name="parse_url",
            content=content or truncate(link.get("cleaned_text", ""), 1200),
            metadata={
                "title": link.get("title", ""),
                "url": link.get("url", url),
                "sources": sources,
                "status": link.get("status", ""),
            },
        )

    def web_search(self, request: ChatRequest) -> ToolResult:
        query = self._search_query_from_message(request.message)
        results, error = self._duckduckgo_search(query)
        if not results:
            results = self._search_local_links(request.user_id, request.chat_id, query)

        content_lines: list[str] = []
        sources: list[str] = []
        for result in results[:5]:
            title = truncate(result.get("title", ""), 120)
            snippet = truncate(result.get("snippet", ""), 240)
            url = result.get("url", "")
            if title and snippet:
                content_lines.append(f"{title}: {snippet}")
            elif title:
                content_lines.append(title)
            elif snippet:
                content_lines.append(snippet)
            if url:
                sources.append(url)

        metadata: dict[str, Any] = {"sources": sources}
        if error and not results:
            metadata["error"] = error

        return ToolResult(
            name="web_search",
            content="\n".join(content_lines),
            metadata=metadata,
        )

    def _ensure_file_attachment_ingested(self, user_id: str, chat_id: str, attachment: Attachment) -> str | None:
        if attachment.file_id and self.storage.get_file(attachment.file_id):
            return attachment.file_id

        file_path = (
            attachment.metadata.get("file_path")
            or attachment.metadata.get("path")
            or attachment.metadata.get("local_path")
        )
        if not file_path:
            return attachment.file_id

        ingestion_result = self.ingestion.ingest_file(
            user_id=user_id,
            chat_id=chat_id,
            file_path=file_path,
            file_id=attachment.file_id,
            mime_type=attachment.mime_type,
            metadata=attachment.metadata,
        )
        return ingestion_result.get("file_id")

    def _resolve_url(self, request: ChatRequest) -> str | None:
        for attachment in request.attachments:
            if attachment.type == "url" and attachment.url:
                return attachment.url
            if attachment.metadata.get("url"):
                return str(attachment.metadata["url"])
        urls = extract_urls(request.message)
        return urls[0] if urls else None

    def _search_query_from_message(self, message: str) -> str:
        query = message.strip()
        for marker in (
            "проверь в интернете",
            "найди",
            "в интернете",
            "в сети",
            "search",
            "latest",
        ):
            query = re.sub(marker, " ", query, flags=re.IGNORECASE)
        query = clean_text(query)
        return query or message

    def _duckduckgo_search(self, query: str) -> tuple[list[dict[str, str]], str | None]:
        try:
            response = requests.get(
                "https://html.duckduckgo.com/html/",
                params={"q": query},
                timeout=10,
                headers={"User-Agent": "MTS-MVP/1.0 (+web-search)"},
            )
            response.raise_for_status()
        except Exception as error:
            return [], str(error)

        soup = BeautifulSoup(response.text, "html.parser")
        results: list[dict[str, str]] = []
        for container in soup.select(".result"):
            title_link = container.select_one("a.result__a")
            snippet_node = container.select_one(".result__snippet")
            if title_link is None:
                continue
            results.append(
                {
                    "title": clean_text(title_link.get_text(" ", strip=True)),
                    "url": title_link.get("href", ""),
                    "snippet": clean_text(snippet_node.get_text(" ", strip=True)) if snippet_node else "",
                }
            )
            if len(results) >= 5:
                break
        return results, None

    def _search_local_links(self, user_id: str, chat_id: str, query: str) -> list[dict[str, str]]:
        links = self.storage.list_links(user_id=user_id)
        top_links = pick_top_rows(
            links,
            query=query,
            text_key="cleaned_text",
            limit=5,
            chat_id=chat_id,
            fallback_if_no_match=True,
        )
        return [
            {
                "title": item.get("title", "") or item.get("url", ""),
                "url": item.get("url", ""),
                "snippet": truncate(item.get("cleaned_text", ""), 240),
            }
            for item in top_links
        ]


class BackendServices:
    def __init__(self, db_path: str | Path | None = None) -> None:
        self.db_path = Path(db_path or DEFAULT_DB_PATH)
        self.storage = StorageBackend(self.db_path)
        self.ingestion = IngestionService(self.storage)
        self.retrieval = RetrievalService(self.storage)
        self.memory = MemoryService(self.storage, self.retrieval)
        self.tools = ToolBackend(self.storage, self.ingestion, self.retrieval)


@lru_cache(maxsize=8)
def get_backend_services(db_path: str | Path | None = None) -> BackendServices:
    resolved = str(Path(db_path or DEFAULT_DB_PATH).resolve())
    return BackendServices(db_path=resolved)


def build_memory_packet(packet: dict[str, Any]) -> MemoryPacket:
    return MemoryPacket(
        chat_summary=packet.get("chat_summary", ""),
        user_facts=list(packet.get("user_facts", [])),
        preferences=list(packet.get("preferences", [])),
        relevant_docs=list(packet.get("relevant_docs", [])),
        ongoing_tasks=list(packet.get("ongoing_tasks", [])),
    )


def attachment_from_dict(payload: dict[str, Any]) -> Attachment:
    return Attachment(
        type=payload.get("type", ""),
        file_id=payload.get("file_id"),
        url=payload.get("url"),
        mime_type=payload.get("mime_type"),
        metadata=dict(payload.get("metadata", {})),
    )


def chat_request_from_dict(payload: dict[str, Any]) -> ChatRequest:
    attachments = [attachment_from_dict(item) for item in payload.get("attachments", [])]
    return ChatRequest(
        user_id=payload["user_id"],
        chat_id=payload["chat_id"],
        message=payload.get("message", ""),
        attachments=attachments,
        mode=payload.get("mode", "auto"),
        selected_model=payload.get("selected_model"),
        task_type=payload.get("task_type"),
    )
